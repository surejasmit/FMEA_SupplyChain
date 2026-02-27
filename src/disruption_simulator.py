"""
Disruption Simulator for FMEA Supply Chain Analysis

Simulates the cascading effect of supply chain node failures on FMEA risk
scores.  Uses Dataset_AI_Supply_Optimization.csv to build a dependency graph
of supply-chain routes and product categories, then applies FMEA risk
escalation rules to every affected component.
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd

logger = logging.getLogger(__name__)


class DisruptionSimulator:
    """
    Simulates supply chain disruptions and their cascading effect on FMEA scores.

    The supply chain dataset (Dataset_AI_Supply_Optimization.csv) is used to
    build a dependency graph where the nodes are Route IDs.  Two routes are
    considered *neighbours* when they both carry at least one shared Product
    Category, modelling the real-world pressure that a failed route places on
    alternative routes for the same product type.

    When a node fails:
    - **Level 1** (direct dependents): FMEA components whose assigned route IS
      the failed node.
    - **Level 2** (indirect dependents): FMEA components on routes that are
      one hop away in the dependency graph (share a Product Category with the
      failed route).

    Escalation rules applied to the FMEA DataFrame
    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    - Level 1: Occurrence +3, Severity +2, Detection +1
    - Level 2: Occurrence +1, Severity +1, Detection +1
    - All scores capped at MAX_SCORE (10)
    - RPN recalculated as Severity x Occurrence x Detection
    - ``Disruption_Delta_RPN`` column added: new RPN - original RPN
    """

    # Escalation constants
    LEVEL1_OCCURRENCE_DELTA: int = 3
    LEVEL1_SEVERITY_DELTA: int = 2
    LEVEL2_OCCURRENCE_DELTA: int = 1
    LEVEL2_SEVERITY_DELTA: int = 1
    ALL_DETECTION_DELTA: int = 1
    MAX_SCORE: int = 10

    def __init__(self, dataset_path: str) -> None:
        """
        Initialise the simulator by loading the supply chain dataset and
        building the dependency graph.

        Args:
            dataset_path: Absolute or relative path to
                          ``Dataset_AI_Supply_Optimization.csv``.

        Raises:
            FileNotFoundError: If *dataset_path* does not exist.
            KeyError: If expected columns are missing from the dataset.
        """
        self.dataset_path: Path = Path(dataset_path)

        # Internal graph structures populated by _build_dependency_graph()
        self._supply_df: pd.DataFrame = pd.DataFrame()
        self._route_to_categories: Dict[int, set] = {}
        self._category_to_routes: Dict[str, set] = {}
        self._route_neighbors: Dict[int, set] = {}
        self._all_routes: List[int] = []

        self._load_dataset()
        self._build_dependency_graph()

    # Public API

    def get_downstream_components(
        self,
        fmea_df: pd.DataFrame,
        failed_node: str,
    ) -> Tuple[List[int], List[int]]:
        """
        Identify which FMEA row indices are affected when the named supply
        chain node fails.

        FMEA components are deterministically assigned to supply chain routes
        by cycling through the available Route IDs (``row_index % num_routes``).
        This preserves a stable, reproducible mapping regardless of FMEA size.

        Args:
            fmea_df:     The current FMEA DataFrame.
            failed_node: Name or identifier of the failed node.  Supported
                         forms:
                         - Route ID: integer string, or prefixed form such as
                           ``"Route_5"``, ``"route 5"``, ``"5"``.
                         - Product Category: e.g. ``"Fresh"``,
                           ``"Refrigerated"`` (case-insensitive).
                         - Traffic condition: ``"High"``, ``"Moderate"``,
                           ``"Low"`` (case-insensitive).

        Returns:
            ``(level1_indices, level2_indices)`` -- disjoint lists of zero-based
            integer row indices from *fmea_df*.
        """
        node_type, node_value = self._parse_failed_node(failed_node)

        level1_routes: set = set()
        level2_routes: set = set()

        if node_type == "route":
            level1_routes = {node_value}
            level2_routes = self._route_neighbors.get(node_value, set()).copy()

        elif node_type == "category":
            level1_routes = self._category_to_routes.get(str(node_value), set()).copy()
            for r in level1_routes:
                level2_routes.update(self._route_neighbors.get(r, set()))
            level2_routes -= level1_routes

        elif node_type == "traffic":
            traffic_col = "Traffic Conditions"
            affected = self._supply_df.loc[
                self._supply_df[traffic_col] == node_value, "Route (ID)"
            ]
            level1_routes = {int(r) for r in affected.unique()}
            for r in level1_routes:
                level2_routes.update(self._route_neighbors.get(r, set()))
            level2_routes -= level1_routes

        else:
            # Unknown node: conservative approach - treat everything as Level 2
            logger.warning(
                "Unknown node '%s'; applying Level-2 impact to all rows.",
                failed_node,
            )
            return [], list(range(len(fmea_df)))

        # Map route sets to FMEA row indices
        level1_indices: List[int] = []
        level2_indices: List[int] = []

        for idx in range(len(fmea_df)):
            assigned_route = self._assign_component_route(idx)
            if assigned_route in level1_routes:
                level1_indices.append(idx)
            elif assigned_route in level2_routes:
                level2_indices.append(idx)

        logger.info(
            "Node '%s' failure -> %d Level-1, %d Level-2 components affected.",
            failed_node,
            len(level1_indices),
            len(level2_indices),
        )
        return level1_indices, level2_indices

    def apply_risk_escalation(
        self,
        fmea_df: pd.DataFrame,
        failed_node: str,
    ) -> pd.DataFrame:
        """
        Apply FMEA risk escalation rules for a node failure.

        Returns a **new** DataFrame (the original is not modified) with updated
        Severity, Occurrence, Detection, and Rpn values, plus a new column
        ``Disruption_Delta_RPN``.

        Args:
            fmea_df:     FMEA DataFrame produced by ``FMEAGenerator``.
            failed_node: Supply chain node identifier to fail.

        Returns:
            Updated DataFrame with ``Disruption_Delta_RPN`` column.
        """
        col_map = self._detect_score_columns(fmea_df)
        sev_col = col_map["severity"]
        occ_col = col_map["occurrence"]
        det_col = col_map["detection"]
        rpn_col = col_map["rpn"]

        result_df = fmea_df.copy()

        # Ensure numeric types; replace non-numeric values with midpoint 5
        for col in (sev_col, occ_col, det_col):
            result_df[col] = pd.to_numeric(result_df[col], errors="coerce").fillna(5)

        # RPN falls back to 0 so original_rpn is never NaN before delta computation
        result_df[rpn_col] = pd.to_numeric(result_df[rpn_col], errors="coerce").fillna(0)

        original_rpn = result_df[rpn_col].copy()

        level1_indices, level2_indices = self.get_downstream_components(
            result_df, failed_node
        )

        # Level 1 escalation
        if level1_indices:
            result_df.loc[level1_indices, occ_col] = (
                result_df.loc[level1_indices, occ_col]
                + self.LEVEL1_OCCURRENCE_DELTA
            ).clip(upper=self.MAX_SCORE)

            result_df.loc[level1_indices, sev_col] = (
                result_df.loc[level1_indices, sev_col]
                + self.LEVEL1_SEVERITY_DELTA
            ).clip(upper=self.MAX_SCORE)

        # Level 2 escalation
        if level2_indices:
            result_df.loc[level2_indices, occ_col] = (
                result_df.loc[level2_indices, occ_col]
                + self.LEVEL2_OCCURRENCE_DELTA
            ).clip(upper=self.MAX_SCORE)

            result_df.loc[level2_indices, sev_col] = (
                result_df.loc[level2_indices, sev_col]
                + self.LEVEL2_SEVERITY_DELTA
            ).clip(upper=self.MAX_SCORE)

        # Detection +1 for all affected rows
        all_affected = level1_indices + level2_indices
        if all_affected:
            result_df.loc[all_affected, det_col] = (
                result_df.loc[all_affected, det_col] + self.ALL_DETECTION_DELTA
            ).clip(upper=self.MAX_SCORE)

        # Recalculate RPN
        result_df[rpn_col] = (
            result_df[sev_col] * result_df[occ_col] * result_df[det_col]
        )

        # Delta column
        result_df["Disruption_Delta_RPN"] = result_df[rpn_col] - original_rpn

        logger.info(
            "Risk escalation complete. Max RPN delta: %.0f",
            result_df["Disruption_Delta_RPN"].max(),
        )
        return result_df

    def export_disruption_report(
        self,
        original_fmea: pd.DataFrame,
        failed_node: str,
        output_path: str,
    ) -> None:
        """
        Write a two-sheet Excel workbook to *output_path*.

        Sheet 1 -- ``Updated_FMEA``
            Full recalculated FMEA table.  Rows where RPN increased after the
            disruption are highlighted in red.

        Sheet 2 -- ``Disruption_Summary``
            - Failed node name
            - Total components affected
            - Level-1 vs Level-2 breakdown
            - Top-3 highest-RPN components after the disruption

        Args:
            original_fmea: Baseline FMEA DataFrame (scores *before* disruption).
            failed_node:   The node whose failure is being simulated.
            output_path:   Destination ``.xlsx`` file path.
        """
        updated_fmea = self.apply_risk_escalation(original_fmea, failed_node)
        level1_indices, level2_indices = self.get_downstream_components(
            original_fmea, failed_node
        )

        col_map = self._detect_score_columns(updated_fmea)
        rpn_col = col_map["rpn"]

        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
            # Sheet 1: Updated FMEA
            updated_fmea.to_excel(writer, sheet_name="Updated_FMEA", index=False)
            self._format_fmea_sheet(
                writer.sheets["Updated_FMEA"],
                updated_fmea,
            )

            # Sheet 2: Disruption Summary
            total_affected = len(level1_indices) + len(level2_indices)

            # Identify the highest-RPN row for the summary
            if len(updated_fmea) > 0 and updated_fmea[rpn_col].notna().any():
                max_rpn_row = updated_fmea.loc[updated_fmea[rpn_col].idxmax()]
                highest_rpn_component = str(
                    max_rpn_row.get(
                        "Component",
                        max_rpn_row.get("component", "N/A"),
                    )
                )
                # Guard against NaN before int conversion
                max_val = updated_fmea[rpn_col].max()
                max_rpn_value = int(max_val) if not pd.isna(max_val) else 0
            else:
                highest_rpn_component = "N/A"
                max_rpn_value = 0

            summary_rows = [
                {"Key": "Failed Node", "Value": failed_node},
                {"Key": "Total Components Affected", "Value": total_affected},
                {"Key": "Level 1 (Direct) Dependents", "Value": len(level1_indices)},
                {"Key": "Level 2 (Indirect) Dependents", "Value": len(level2_indices)},
                {"Key": "Highest RPN Component", "Value": highest_rpn_component},
                {"Key": "Maximum RPN After Disruption", "Value": max_rpn_value},
            ]
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(
                writer,
                sheet_name="Disruption_Summary",
                index=False,
                startrow=0,
            )

            # Top-3 table written below the summary block
            top3 = self._build_top3_table(updated_fmea, rpn_col)
            top3_start_row = len(summary_rows) + 2  # zero-based offset

            ws_sum = writer.sheets["Disruption_Summary"]
            self._write_top3_table(ws_sum, top3, top3_start_row)
            self._format_summary_sheet(ws_sum)

        logger.info("Disruption report written to: %s", out_path)

    # Private helpers

    def _load_dataset(self) -> None:
        """Load the CSV and normalise column names.

        The dataset header contains a degree symbol (in 'Average Temperature')
        which is encoded as a Latin-1 byte.  We try UTF-8 first for
        forward-compatibility and fall back to ``latin1`` automatically if
        decoding fails.
        """
        if not self.dataset_path.exists():
            raise FileNotFoundError(
                f"Supply chain dataset not found at: {self.dataset_path}"
            )
        try:
            self._supply_df = pd.read_csv(self.dataset_path, encoding="utf-8")
        except UnicodeDecodeError:
            self._supply_df = pd.read_csv(self.dataset_path, encoding="latin1")
        self._supply_df.columns = [c.strip() for c in self._supply_df.columns]
        logger.info(
            "Loaded supply chain dataset: %d rows, %d columns.",
            len(self._supply_df),
            len(self._supply_df.columns),
        )

    def _build_dependency_graph(self) -> None:
        """
        Build route-level dependency structures from the supply chain dataset.

        Two routes become *neighbours* when they both carry at least one shared
        Product Category, capturing the substitution dependency in real logistics
        (if Route A closes, its product types must divert to routes that also
        carry those product types).
        """
        route_col = "Route (ID)"
        category_col = "Product Category"

        for required in (route_col, category_col):
            if required not in self._supply_df.columns:
                raise KeyError(
                    f"Required column '{required}' not found in dataset. "
                    f"Available: {list(self._supply_df.columns)}"
                )

        # Route -> set of Product Categories
        for route_id, group in self._supply_df.groupby(route_col):
            self._route_to_categories[int(route_id)] = set(
                group[category_col].dropna().unique()
            )

        # Category -> set of Route IDs
        for route_id, cats in self._route_to_categories.items():
            for cat in cats:
                self._category_to_routes.setdefault(cat, set()).add(route_id)

        # Route adjacency (shared product category)
        self._all_routes = sorted(self._route_to_categories.keys())
        for route_id in self._all_routes:
            neighbours: set = set()
            for cat in self._route_to_categories[route_id]:
                neighbours.update(self._category_to_routes[cat])
            neighbours.discard(route_id)
            self._route_neighbors[route_id] = neighbours

        logger.info(
            "Dependency graph built: %d routes, categories: %s.",
            len(self._all_routes),
            sorted(self._category_to_routes.keys()),
        )

    def _assign_component_route(self, component_index: int) -> int:
        """
        Deterministically assign an FMEA row (by its zero-based index) to a
        supply chain route by cycling through the sorted list of Route IDs.

        Args:
            component_index: Zero-based row index.

        Returns:
            Route ID for this component.
        """
        num_routes = len(self._all_routes)
        if num_routes == 0:
            return 0
        return self._all_routes[component_index % num_routes]

    def _parse_failed_node(self, failed_node: str) -> Tuple[str, object]:
        """
        Classify *failed_node* into ``("route", id)``, ``("category", name)``,
        ``("traffic", value)``, or ``("unknown", raw_string)``.

        Matching order:
        1. Known Product Category (case-insensitive)
        2. Known Traffic Conditions value (case-insensitive)
        3. Numeric Route ID (extracted with regex)
        4. Unknown fallback

        Returns:
            ``(node_type, node_value)``
        """
        node_clean = failed_node.strip()
        node_lower = node_clean.lower()

        # 1. Product Category
        known_cats = {c.lower(): c for c in self._category_to_routes}
        if node_lower in known_cats:
            return "category", known_cats[node_lower]

        # 2. Traffic Conditions
        traffic_col = "Traffic Conditions"
        if traffic_col in self._supply_df.columns:
            known_traffic = {
                v.lower(): v
                for v in self._supply_df[traffic_col].dropna().unique()
            }
            if node_lower in known_traffic:
                return "traffic", known_traffic[node_lower]

        # 3. Numeric Route ID
        match = re.search(r"\d+", node_clean)
        if match:
            route_id = int(match.group())
            if route_id in self._route_to_categories:
                return "route", route_id

        logger.warning(
            "Could not classify node '%s' as a known route, category, or "
            "traffic condition.",
            failed_node,
        )
        return "unknown", node_clean

    def _detect_score_columns(self, fmea_df: pd.DataFrame) -> Dict[str, str]:
        """
        Return a ``{canonical_name: actual_column_name}`` mapping for the four
        FMEA numeric score columns, tolerating both Title Case (output of
        ``FMEAGenerator._format_output``) and lowercase variants.

        Raises:
            KeyError: If a required column cannot be found.
        """
        candidates: Dict[str, List[str]] = {
            "severity":   ["Severity",   "severity"],
            "occurrence": ["Occurrence", "occurrence"],
            "detection":  ["Detection",  "detection"],
            "rpn":        ["Rpn",        "rpn",       "RPN"],
        }
        result: Dict[str, str] = {}
        for key, names in candidates.items():
            for name in names:
                if name in fmea_df.columns:
                    result[key] = name
                    break
            if key not in result:
                raise KeyError(
                    f"Could not locate '{key}' column in FMEA DataFrame. "
                    f"Available columns: {list(fmea_df.columns)}"
                )
        return result

    def _build_top3_table(
        self, updated_fmea: pd.DataFrame, rpn_col: str
    ) -> pd.DataFrame:
        """Return the top-3 rows by post-disruption RPN with relevant columns."""
        display_candidates = [
            "Component", "component", "Failure Mode", "failure_mode",
            rpn_col, "Disruption_Delta_RPN",
        ]
        display_cols = [c for c in display_candidates if c in updated_fmea.columns]
        top3 = (
            updated_fmea.nlargest(3, rpn_col)[display_cols]
            .reset_index(drop=True)
        )
        top3.index = top3.index + 1
        return top3

    def _format_fmea_sheet(self, ws, updated_fmea: pd.DataFrame) -> None:
        """Apply header styling and red-fill to rows where RPN increased."""
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        red_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        delta_col = "Disruption_Delta_RPN"
        if delta_col in updated_fmea.columns:
            for row_num, delta in enumerate(updated_fmea[delta_col].values, start=2):
                if pd.notna(delta) and delta > 0:
                    for cell in ws[row_num]:
                        cell.fill = red_fill

        for col_idx, col_name in enumerate(updated_fmea.columns, start=1):
            col_values = updated_fmea[col_name].astype(str)
            # Guard against NaN from str.len().max() when all values are NaN
            val_max = col_values.str.len().fillna(0).max() if len(col_values) > 0 else 0
            val_max = 0 if pd.isna(val_max) else int(val_max)
            max_len = max(len(str(col_name)), val_max)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 4, 50
            )

    def _format_summary_sheet(self, ws) -> None:
        """Bold and blue-background the header row of the summary sheet."""
        from openpyxl.styles import Font, PatternFill

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    def _write_top3_table(self, ws, top3: pd.DataFrame, start_row: int) -> None:
        """Write the top-3 table into the summary worksheet below the KPI block."""
        from openpyxl.styles import Font

        # Section heading (1-based row in openpyxl = start_row + 1)
        heading_row = start_row + 1
        ws.cell(row=heading_row, column=1).value = (
            "Top-3 Components by RPN After Disruption"
        )
        ws.cell(row=heading_row, column=1).font = Font(bold=True)

        # Column headers
        col_headers = ["Rank"] + list(top3.columns)
        header_row = heading_row + 1
        for col_idx, header in enumerate(col_headers, start=1):
            ws.cell(row=header_row, column=col_idx).value = header

        # Data rows
        for rank, (_, row_data) in enumerate(top3.iterrows(), start=1):
            data_row = header_row + rank
            ws.cell(row=data_row, column=1).value = rank
            for col_idx, value in enumerate(row_data.values, start=2):
                ws.cell(row=data_row, column=col_idx).value = (
                    int(value) if isinstance(value, float) and value.is_integer()
                    else value
                )
